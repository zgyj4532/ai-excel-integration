import pandas as pd
import openpyxl
from openpyxl import load_workbook

def analyze_excel_file(filename):
    """分析Excel文件以检查其内容"""
    print(f"\n分析文件: {filename}")
    
    # 使用pandas读取Excel
    df = pd.read_excel(filename)
    print("DataFrame内容:")
    print(df)
    print(f"列名: {list(df.columns)}")
    print(f"形状: {df.shape}")
    
    # 使用openpyxl直接读取单元格值
    print("\n使用openpyxl直接读取:")
    workbook = load_workbook(filename)
    sheet = workbook.active
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        print(f"行 {row_idx}: {row}")
        if row_idx >= 6:  # 只打印前6行
            break

# 分析薪资估算文件
analyze_excel_file('modified_salary_file.xlsx')
print("\n" + "="*50)

# 分析部门文件
analyze_excel_file('modified_file.xlsx')