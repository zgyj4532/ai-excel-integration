import openpyxl
from openpyxl.utils import get_column_letter

def analyze_excel_file(filename):
    """分析Excel文件单元格内容和类型"""
    print(f"分析文件: {filename}")
    
    # 加载工作簿
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    
    print(f"工作表名称: {sheet.title}")
    print(f"最大行数: {sheet.max_row}, 最大列数: {sheet.max_column}")
    
    # 遍历所有单元格
    for row_idx in range(1, sheet.max_row + 1):
        row_data = []
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
            cell_value = cell.value
            cell_type = cell.data_type  # 获取单元格数据类型
            formula = cell.formula if hasattr(cell, 'formula') and cell.formula else None
            
            cell_info = f"{cell_ref}: val={cell_value}, type={cell_type}"
            if formula:
                cell_info += f", formula={formula}"
            
            row_data.append(cell_info)
        
        print(f"  行 {row_idx}: {' | '.join(row_data)}")

if __name__ == "__main__":
    analyze_excel_file("final_test_success.xlsx")